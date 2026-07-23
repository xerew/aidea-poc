from io import BytesIO

from django.contrib.auth.models import User
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from hub.models import (
    Course,
    Enrollment,
    LearningPillar,
    Lesson,
    LessonProgress,
    Module,
    UserProfile,
)


def make_creator(username='det_cc'):
    u = User.objects.create_user(username=username, password='pass12345')
    UserProfile.objects.create(user=u, user_type=UserProfile.UserType.CONTENT_CREATOR)
    return u


def make_teacher(username, first='', last=''):
    u = User.objects.create_user(username=username, password='pass12345', first_name=first, last_name=last)
    UserProfile.objects.create(user=u, user_type=UserProfile.UserType.TEACHER)
    return u


class TeacherDetailTests(APITestCase):
    def setUp(self):
        self.creator = make_creator()
        self.other = make_creator('det_other')
        self.pillar = LearningPillar.objects.create(name='P', slug='p-det', order=1)
        self.course = Course.objects.create(
            title='Detail Course', pillar=self.pillar, level='beginner',
            duration_hours=1, is_published=True, created_by=self.creator,
        )
        self.module = Module.objects.create(title='M', course=self.course, order=1)
        self.quiz = Lesson.objects.create(
            module=self.module, title='Quiz 1', lesson_type='quiz', order=1, is_required=True,
            quiz_data=[{
                'question': 'What is 2+2?',
                'options': [
                    {'text': '3', 'is_correct': False},
                    {'text': '4', 'is_correct': True},
                ],
            }],
        )
        self.teacher = make_teacher('det_t1', 'Ada', 'Byte')
        Enrollment.objects.create(user=self.teacher, course=self.course, progress_pct=100)
        LessonProgress.objects.create(
            user=self.teacher, lesson=self.quiz,
            time_spent_seconds=120, quiz_score=1.0,
            quiz_answers=[True],
            engagement_data={'quiz_selected': [1]},
        )
        self.url = reverse('analytics-course-teachers', kwargs={'pk': self.course.pk})

    def test_teacher_role_forbidden(self):
        self.client.force_authenticate(self.teacher)
        self.assertEqual(self.client.get(self.url).status_code, status.HTTP_403_FORBIDDEN)

    def test_non_owner_creator_gets_404(self):
        self.client.force_authenticate(self.other)
        self.assertEqual(self.client.get(self.url).status_code, status.HTTP_404_NOT_FOUND)

    def test_owner_sees_teacher_detail_with_answers(self):
        self.client.force_authenticate(self.creator)
        res = self.client.get(self.url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(len(res.data['teachers']), 1)
        teacher = res.data['teachers'][0]
        self.assertEqual(teacher['name'], 'Ada Byte')
        self.assertEqual(teacher['time_spent_seconds'], 120)
        self.assertEqual(teacher['progress_pct'], 100)
        q = teacher['quizzes'][0]['questions'][0]
        self.assertEqual(q['selected_text'], '4')
        self.assertEqual(q['correct_text'], '4')
        self.assertTrue(q['is_correct'])


class ExportTests(APITestCase):
    def setUp(self):
        self.creator = make_creator('exp_cc')
        self.teacher = make_teacher('exp_t1', 'Bo', 'Bit')
        self.pillar = LearningPillar.objects.create(name='P', slug='p-exp', order=1)
        self.course = Course.objects.create(
            title='Export Course', pillar=self.pillar, level='beginner',
            duration_hours=1, is_published=True, created_by=self.creator,
        )
        Enrollment.objects.create(user=self.teacher, course=self.course, progress_pct=50)
        self.url = reverse('analytics-export')

    def test_teacher_forbidden(self):
        self.client.force_authenticate(self.teacher)
        self.assertEqual(self.client.get(self.url).status_code, status.HTTP_403_FORBIDDEN)

    def test_export_has_sheet_per_course_with_teacher_rows(self):
        self.client.force_authenticate(self.creator)
        res = self.client.get(self.url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', res['Content-Type'])
        wb = load_workbook(BytesIO(res.getvalue()))
        self.assertIn('Export Course', wb.sheetnames)
        ws = wb['Export Course']
        self.assertEqual(ws['A1'].value, 'Teacher')
        self.assertEqual(ws['A2'].value, 'Bo Bit')
        self.assertEqual(ws['D2'].value, 50)  # progress %

    def test_export_empty_when_no_authored_courses(self):
        creator2 = make_creator('exp_cc2')
        self.client.force_authenticate(creator2)
        res = self.client.get(self.url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        wb = load_workbook(BytesIO(res.getvalue()))
        self.assertEqual(wb.sheetnames, ['No courses'])
