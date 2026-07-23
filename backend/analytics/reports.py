"""Per-teacher analytics: drill-down report (#27) and xlsx export (#26).

Both are scoped by the caller to a content creator's own authored courses.
All the underlying data already lives on LessonProgress — time_spent_seconds,
quiz_score, quiz_answers (per-question correct/incorrect) and
engagement_data['quiz_selected'] (the option index the teacher picked).
"""
import re

from hub.models import Enrollment, Lesson, LessonProgress


def _quiz_questions(lesson, lp):
    """Reconstruct a teacher's answers for one quiz lesson."""
    selected = (lp.engagement_data or {}).get('quiz_selected', [])
    booleans = lp.quiz_answers or []
    questions = []
    for i, q in enumerate(lesson.quiz_data or []):
        options = q.get('options', [])
        sel_idx = selected[i] if i < len(selected) and isinstance(selected[i], int) else None
        selected_text = (
            options[sel_idx].get('text')
            if sel_idx is not None and 0 <= sel_idx < len(options)
            else None
        )
        correct_text = next((o.get('text') for o in options if o.get('is_correct')), None)
        questions.append({
            'question': q.get('question', ''),
            'selected_text': selected_text,
            'is_correct': booleans[i] if i < len(booleans) else None,
            'correct_text': correct_text,
        })
    return questions


def _teacher_rows(course):
    """Yield a per-teacher summary + quiz detail dict for each enrollment."""
    quiz_lessons = list(
        Lesson.objects.filter(module__course=course, lesson_type='quiz')
        .order_by('module__order', 'order')
    )
    enrollments = (
        Enrollment.objects.filter(course=course)
        .select_related('user')
        .order_by('user__first_name', 'user__last_name', 'user__username')
    )
    for enrollment in enrollments:
        user = enrollment.user
        progresses = {
            lp.lesson_id: lp
            for lp in LessonProgress.objects.filter(user=user, lesson__module__course=course)
        }
        time_spent = sum((lp.time_spent_seconds or 0) for lp in progresses.values())

        quizzes, scores = [], []
        for lesson in quiz_lessons:
            lp = progresses.get(lesson.id)
            if lp is None:
                continue
            if lp.quiz_score is not None:
                scores.append(lp.quiz_score)
            quizzes.append({
                'lesson_id': lesson.id,
                'lesson_title': lesson.title,
                'score': lp.quiz_score,
                'questions': _quiz_questions(lesson, lp),
            })

        avg_quiz_score = round(sum(scores) / len(scores), 2) if scores else None
        yield {
            'user_id': user.id,
            'name': user.get_full_name() or user.username,
            'email': user.email,
            'enrolled_at': enrollment.enrolled_at,
            'progress_pct': enrollment.progress_pct,
            'completed': enrollment.progress_pct == 100,
            'time_spent_seconds': time_spent,
            'avg_quiz_score': avg_quiz_score,
            'quizzes': quizzes,
        }


def build_course_teacher_report(course):
    """#27 — full per-teacher detail (with quiz answers) for one course."""
    return {
        'course_id': course.id,
        'course_title': course.title,
        'teachers': list(_teacher_rows(course)),
    }


# ── Excel export (#26) ─────────────────────────────────────────────────────────

_SHEET_HEADERS = [
    'Teacher', 'Email', 'Enrolled', 'Progress %', 'Completed',
    'Time spent (min)', 'Avg quiz score %', 'Quizzes taken',
]
_INVALID_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')


def _sheet_name(title, used):
    """Excel sheet names: <=31 chars, no []:*?/\\, unique within the workbook."""
    name = _INVALID_SHEET_CHARS.sub(' ', title).strip()[:31] or 'Course'
    candidate, n = name, 2
    while candidate.lower() in used:
        suffix = f' ({n})'
        candidate = name[:31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def build_analytics_workbook(courses):
    """One sheet per course, one row per enrolled teacher."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    for course in courses:
        ws = wb.create_sheet(_sheet_name(course.title, used_names))
        for col, header in enumerate(_SHEET_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        for row in _teacher_rows(course):
            ws.append([
                row['name'],
                row['email'],
                row['enrolled_at'].strftime('%Y-%m-%d') if row['enrolled_at'] else '',
                row['progress_pct'],
                'yes' if row['completed'] else 'no',
                round(row['time_spent_seconds'] / 60, 1),
                round(row['avg_quiz_score'] * 100) if row['avg_quiz_score'] is not None else '',
                len(row['quizzes']),
            ])
        for col in range(1, len(_SHEET_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    if not wb.sheetnames:
        ws = wb.create_sheet('No courses')
        ws['A1'] = 'You have not authored any courses yet.'

    return wb
