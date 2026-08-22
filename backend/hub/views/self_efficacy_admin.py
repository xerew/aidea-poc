import csv
from collections import defaultdict
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework.views import APIView

from hub.models import OnboardingDimension, SelfEfficacyAttempt, SelfEfficacyConfig
from hub.psychometrics import compute_scale_reliability
from hub.views.permissions import IsAdmin


def _active_dimensions():
    return list(
        OnboardingDimension.objects.filter(is_active=True)
        .order_by('order').prefetch_related('questions')
    )


def _first_attempt_per_user():
    """One SelfEfficacyAttempt per user — their earliest (baseline) — so a
    reliability/validation analysis uses independent observations."""
    seen, firsts = set(), []
    for attempt in (
        SelfEfficacyAttempt.objects
        .select_related('user__profile__subject')
        .order_by('user_id', 'created_at')
    ):
        if attempt.user_id in seen:
            continue
        seen.add(attempt.user_id)
        firsts.append(attempt)
    return firsts

# Self-efficacy bands are shown to users as competency levels.
BAND_LABELS = {'low': 'Beginner', 'moderate': 'Intermediate', 'high': 'Advanced'}


class AdminSelfEfficacyView(APIView):
    """GET/PATCH the retake switch. Opening it (off → on) starts a new window,
    recorded in `retake_opened_at`; each teacher may retake once per window."""
    permission_classes = [IsAdmin]

    def _payload(self, cfg):
        return {
            'retake_open': cfg.retake_open,
            'retake_opened_at': cfg.retake_opened_at.isoformat() if cfg.retake_opened_at else None,
        }

    def get(self, request):
        return Response(self._payload(SelfEfficacyConfig.get()))

    def patch(self, request):
        cfg = SelfEfficacyConfig.get()
        new_open = bool(request.data.get('retake_open'))
        if new_open and not cfg.retake_open:
            cfg.retake_opened_at = timezone.now()  # a fresh window
        cfg.retake_open = new_open
        cfg.save(update_fields=['retake_open', 'retake_opened_at'])
        return Response(self._payload(cfg))


class AdminSelfEfficacyExportView(APIView):
    """GET → every teacher's answers and scores as an XLSX. Each retake lands on
    its own sheet ('AI Comp Version 1', 'AI Comp Version 2', …) so admins can
    compare rounds. A 'Questions' sheet maps the Q-codes to their text."""
    permission_classes = [IsAdmin]

    def get(self, request):
        dimensions = list(
            OnboardingDimension.objects.filter(is_active=True)
            .order_by('order').prefetch_related('questions')
        )
        questions = []  # ordered [(dimension, question), …]
        for dim in dimensions:
            for q in sorted((q for q in dim.questions.all() if q.is_active), key=lambda q: q.order):
                questions.append((dim, q))
        qcodes = [f'Q{i + 1}' for i in range(len(questions))]

        header = (
            ['User ID', 'Name', 'Username', 'Email', 'Date']
            + qcodes
            + [f'{dim.name} avg' for dim in dimensions]
            + ['Overall avg', 'Assessment', 'Competency score (0-6)']
        )

        # Group each teacher's attempts chronologically → attempt N = version N.
        by_user = defaultdict(list)
        for a in SelfEfficacyAttempt.objects.select_related('user').order_by('user_id', 'created_at'):
            by_user[a.user_id].append(a)
        max_versions = max((len(v) for v in by_user.values()), default=0)

        def row(attempt):
            u = attempt.user
            name = f'{u.first_name} {u.last_name}'.strip() or u.username
            answers = attempt.answers or {}
            dims = attempt.dimension_scores or {}
            return (
                [u.id, name, u.username, u.email,
                 timezone.localtime(attempt.created_at).strftime('%Y-%m-%d %H:%M')]
                + [answers.get(str(q.id), '') for (_d, q) in questions]
                + [dims.get(dim.slug, {}).get('average', '') for dim in dimensions]
                + [attempt.overall_average,
                   BAND_LABELS.get(attempt.overall_band, attempt.overall_band),
                   attempt.competency_score]
            )

        wb = Workbook()
        wb.remove(wb.active)
        for v in range(max(max_versions, 1)):
            ws = wb.create_sheet(title=f'AI Comp Version {v + 1}')
            ws.append(header)
            entries = sorted(
                (alist for alist in by_user.values() if len(alist) > v),
                key=lambda alist: alist[0].user.username,
            )
            for alist in entries:
                ws.append(row(alist[v]))

        ref = wb.create_sheet(title='Questions')
        ref.append(['Code', 'Dimension', 'Question'])
        for code, (dim, q) in zip(qcodes, questions):
            ref.append([code, dim.name, q.text])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="aidea-ai-competency.xlsx"'
        return response


class AdminSelfEfficacyPsychometricsView(APIView):
    """GET → classical-test-theory reliability of the self-efficacy scale
    (Cronbach's alpha per dimension + overall, item stats, inter-dimension
    correlations), computed over each teacher's first completed attempt."""
    permission_classes = [IsAdmin]

    def get(self, request):
        dimensions = _active_dimensions()
        responses = [
            {int(k): v for k, v in (a.answers or {}).items()}
            for a in _first_attempt_per_user()
        ]
        return Response(compute_scale_reliability(responses, dimensions))


class AdminSelfEfficacyResearchExportView(APIView):
    """GET → research-ready CSV: one pseudonymous row per participant (first
    attempt) with demographics, the 24 item scores, the six dimension means and
    the overall mean — a wide data matrix for R/SPSS (reliability, EFA/CFA)."""
    permission_classes = [IsAdmin]

    def get(self, request):
        dimensions = _active_dimensions()
        dim_questions = [
            (dim, sorted((q for q in dim.questions.all() if q.is_active), key=lambda q: q.order))
            for dim in dimensions
        ]
        flat = [q for _dim, qs in dim_questions for q in qs]
        qcodes = [f'Q{i}' for i in range(1, len(flat) + 1)]

        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            ['participant', 'subject', 'teaching_level', 'country', 'language', 'date']
            + qcodes
            + [f'{dim.slug}_mean' for dim in dimensions]
            + ['overall_mean']
        )

        def mean(values):
            return round(sum(values) / len(values), 3) if values else ''

        for idx, attempt in enumerate(_first_attempt_per_user(), start=1):
            profile = attempt.user.profile
            answers = {int(k): v for k, v in (attempt.answers or {}).items()}
            row = [
                f'P{idx:04d}',
                profile.subject.slug if profile.subject else '',
                profile.teaching_level or '',
                profile.country or '',
                profile.language,
                attempt.created_at.strftime('%Y-%m-%d'),
            ]
            row += [answers.get(q.id, '') for q in flat]
            row += [mean([answers[q.id] for q in qs if q.id in answers]) for _dim, qs in dim_questions]
            row.append(mean([answers[q.id] for q in flat if q.id in answers]))
            writer.writerow(row)

        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="aidea-self-efficacy-research.csv"'
        return response
