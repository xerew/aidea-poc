"""Course <-> xlsx workbook conversion. One workbook = one course."""
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from hub.models import Course, LearningPillar, Lesson

COURSE_HEADERS = ['title', 'description', 'pillar_slug', 'level',
                  'duration_hours', 'content_format', 'learning_outcomes']
MODULE_HEADERS = ['order', 'title', 'description', 'duration_minutes']
LESSON_HEADERS = ['module_order', 'order', 'title', 'description',
                  'lesson_type', 'content', 'duration_minutes', 'required']
QUIZ_HEADERS   = ['module_order', 'lesson_order', 'question_order', 'question',
                  'option_a', 'option_b', 'option_c', 'option_d', 'option_e',
                  'option_f', 'correct']
OPTION_LETTERS = ['A', 'B', 'C', 'D', 'E', 'F']
DROPDOWN_ROWS  = 500

README_LINES = [
    'AIDEA course workbook',
    '',
    'This file describes ONE course. Import it in Authoring -> Import course.',
    'Importing always creates a NEW unpublished draft owned by you.',
    '',
    'Sheets:',
    '  Course  - exactly one row (row 2). pillar_slug, level and content_format',
    '            offer dropdowns. learning_outcomes: one outcome per line in the cell',
    '            (Alt+Enter inside Excel).',
    '  Modules - one row per module. "order" must be a unique positive number.',
    '  Lessons - one row per lesson. module_order refers to the Modules sheet.',
    '            lesson_type and required offer dropdowns.',
    '  Quiz    - one row per question, only for lessons whose type is quiz.',
    '            module_order and lesson_order refer to the Lessons sheet.',
    '            Fill option_a..option_f (at least two) and put the correct',
    '            letter(s) in "correct", comma separated, e.g. B or A,C.',
    '',
    'Do not rename sheets or reorder columns. The hidden Choices sheet feeds',
    'the dropdowns - leave it alone.',
]


def _write_headers(ws, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)


def _list_validation(wb, ws, choices_col, n_choices, target_col, last_row):
    dv = DataValidation(
        type='list',
        formula1=f'Choices!${choices_col}$1:${choices_col}${n_choices}',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f'{target_col}2:{target_col}{last_row}')


def build_course_workbook(course: Course | None = None) -> Workbook:
    """Build a one-course workbook. With ``course=None`` the sheets, headers,
    dropdowns and README are produced with no data rows — a blank import
    template."""
    wb = Workbook()

    readme = wb.active
    readme.title = 'README'
    for i, line in enumerate(README_LINES, start=1):
        readme.cell(row=i, column=1, value=line)
    readme.column_dimensions['A'].width = 90

    # ── Choices (hidden) ────────────────────────────────────────────────
    choices = wb.create_sheet('Choices')
    levels          = [c[0] for c in Course.Level.choices]
    content_formats = [c[0] for c in Course.ContentFormat.choices]
    lesson_types    = [c[0] for c in Lesson.LessonType.choices]
    pillar_slugs    = list(LearningPillar.objects.values_list('slug', flat=True))
    yes_no          = ['yes', 'no']
    for col, values in enumerate([levels, content_formats, lesson_types, pillar_slugs, yes_no], start=1):
        for row, value in enumerate(values, start=1):
            choices.cell(row=row, column=col, value=value)
    choices.sheet_state = 'hidden'

    # ── Course ──────────────────────────────────────────────────────────
    course_ws = wb.create_sheet('Course')
    _write_headers(course_ws, COURSE_HEADERS)
    if course is not None:
        course_ws.append([
            course.title,
            course.description,
            course.pillar.slug,
            course.level,
            course.duration_hours,
            course.content_format,
            '\n'.join(course.learning_outcomes or []),
        ])
        course_ws['G2'].alignment = course_ws['G2'].alignment.copy(wrap_text=True)
    _list_validation(wb, course_ws, 'D', max(len(pillar_slugs), 1), 'C', 2)
    _list_validation(wb, course_ws, 'A', len(levels), 'D', 2)
    _list_validation(wb, course_ws, 'B', len(content_formats), 'F', 2)

    # ── Modules ─────────────────────────────────────────────────────────
    modules_ws = wb.create_sheet('Modules')
    _write_headers(modules_ws, MODULE_HEADERS)
    modules = list(course.modules.order_by('order').prefetch_related('lessons')) if course else []
    for module in modules:
        modules_ws.append([module.order, module.title, module.description, module.duration_minutes])

    # ── Lessons ─────────────────────────────────────────────────────────
    lessons_ws = wb.create_sheet('Lessons')
    _write_headers(lessons_ws, LESSON_HEADERS)
    quiz_rows = []
    for module in modules:
        for lesson in module.lessons.order_by('order'):
            lessons_ws.append([
                module.order,
                lesson.order,
                lesson.title,
                lesson.description,
                lesson.lesson_type,
                lesson.content,
                lesson.duration_minutes,
                'yes' if lesson.is_required else 'no',
            ])
            if lesson.lesson_type == Lesson.LessonType.QUIZ:
                for q_idx, question in enumerate(lesson.quiz_data or [], start=1):
                    options = question.get('options', [])[:len(OPTION_LETTERS)]
                    texts = [opt.get('text', '') for opt in options]
                    texts += [''] * (len(OPTION_LETTERS) - len(texts))
                    correct = ','.join(
                        OPTION_LETTERS[i] for i, opt in enumerate(options)
                        if opt.get('is_correct')
                    )
                    quiz_rows.append([
                        module.order, lesson.order, q_idx,
                        question.get('question', ''), *texts, correct,
                    ])
    _list_validation(wb, lessons_ws, 'C', len(lesson_types), 'E', DROPDOWN_ROWS)
    _list_validation(wb, lessons_ws, 'E', len(yes_no), 'H', DROPDOWN_ROWS)

    # ── Quiz ────────────────────────────────────────────────────────────
    quiz_ws = wb.create_sheet('Quiz')
    _write_headers(quiz_ws, QUIZ_HEADERS)
    for row in quiz_rows:
        quiz_ws.append(row)

    for ws in (course_ws, modules_ws, lessons_ws, quiz_ws):
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    return wb


MAX_IMPORT_BYTES = 5 * 1024 * 1024

_YES_NO = {'yes': True, 'no': False, '': True, None: True}


def _cell(sheet, col_idx, row):
    return f'{sheet}!{get_column_letter(col_idx)}{row}'


def _rows(ws):
    """Non-empty data rows as (row_number, values) with header-length padding."""
    for row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(v not in (None, '') for v in values):
            yield row_num, values


# PositiveSmallIntegerField maps to Postgres smallint: CHECK (>= 0), max 32767.
# Values outside this range must fail phase-1 validation, not blow up the insert.
INT_MAX = 32767


def _as_int(value, default=0):
    if value in (None, ''):
        return default, True
    if isinstance(value, float) and not value.is_integer():
        return default, False
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default, False
    if not 0 <= number <= INT_MAX:
        return default, False
    return number, True


def parse_course_workbook(file):  # noqa: C901 - single cohesive validator
    from openpyxl import load_workbook

    errors: list[str] = []
    try:
        wb = load_workbook(file, data_only=True)
    except Exception:
        return None, ['File is not a valid xlsx workbook.']

    for sheet in ('Course', 'Modules', 'Lessons'):
        if sheet not in wb.sheetnames:
            errors.append(f'Missing sheet: {sheet}.')
    if errors:
        return None, errors

    levels          = {c[0] for c in Course.Level.choices}
    content_formats = {c[0] for c in Course.ContentFormat.choices}
    lesson_types    = {c[0] for c in Lesson.LessonType.choices}
    pillar_by_slug  = {p.slug: p for p in LearningPillar.objects.all()}

    # ── Course sheet ────────────────────────────────────────────────────
    course_rows = list(_rows(wb['Course']))
    if not course_rows:
        return None, ['Course!A2: course row is missing.']
    row_num, values = course_rows[0]
    values = list(values) + [None] * (len(COURSE_HEADERS) - len(values))
    title, description, pillar_slug, level, duration_hours, content_format, outcomes = values[:7]

    if not (title or '').strip():
        errors.append(f'{_cell("Course", 1, row_num)}: title is required.')
    if pillar_slug not in pillar_by_slug:
        errors.append(f'{_cell("Course", 3, row_num)}: unknown pillar_slug {pillar_slug!r}.')
    if level not in levels:
        errors.append(f'{_cell("Course", 4, row_num)}: level must be one of {sorted(levels)}.')
    duration_hours, ok = _as_int(duration_hours)
    if not ok:
        errors.append(f'{_cell("Course", 5, row_num)}: duration_hours must be a whole number between 0 and 32767.')
    if content_format in (None, ''):
        content_format = Course.ContentFormat.MIXED
    elif content_format not in content_formats:
        errors.append(f'{_cell("Course", 6, row_num)}: content_format must be one of {sorted(content_formats)}.')

    course_payload = {
        'title': (title or '').strip(),
        'description': description or '',
        'pillar': pillar_by_slug.get(pillar_slug),
        'level': level,
        'duration_hours': duration_hours,
        'content_format': content_format,
        'learning_outcomes': [
            line.strip() for line in str(outcomes or '').splitlines() if line.strip()
        ],
    }

    # ── Modules sheet ───────────────────────────────────────────────────
    modules: dict[int, dict] = {}
    for row_num, values in _rows(wb['Modules']):
        values = list(values) + [None] * (len(MODULE_HEADERS) - len(values))
        order, m_title, m_desc, m_minutes = values[:4]
        order, ok = _as_int(order, default=-1)
        if not ok or order < 1:
            errors.append(f'{_cell("Modules", 1, row_num)}: order must be a positive number.')
            continue
        if order in modules:
            errors.append(f'{_cell("Modules", 1, row_num)}: duplicate module order {order}.')
            continue
        if not (m_title or '').strip():
            errors.append(f'{_cell("Modules", 2, row_num)}: title is required.')
            continue
        m_minutes, ok = _as_int(m_minutes)
        if not ok:
            errors.append(f'{_cell("Modules", 4, row_num)}: duration_minutes must be a whole number between 0 and 32767.')
        modules[order] = {
            'order': order, 'title': m_title.strip(), 'description': m_desc or '',
            'duration_minutes': m_minutes, 'lessons': {},
        }
    if not modules:
        errors.append('Modules!A2: at least one module is required.')

    # ── Lessons sheet ───────────────────────────────────────────────────
    for row_num, values in _rows(wb['Lessons']):
        values = list(values) + [None] * (len(LESSON_HEADERS) - len(values))
        m_order, order, l_title, l_desc, l_type, content, minutes, required = values[:8]
        m_order, ok = _as_int(m_order, default=-1)
        if not ok or m_order not in modules:
            errors.append(f'{_cell("Lessons", 1, row_num)}: module_order {m_order!r} does not match any module.')
            continue
        order, ok = _as_int(order, default=-1)
        if not ok or order < 1:
            errors.append(f'{_cell("Lessons", 2, row_num)}: order must be a positive number.')
            continue
        if order in modules[m_order]['lessons']:
            errors.append(f'{_cell("Lessons", 2, row_num)}: duplicate lesson order {order} in module {m_order}.')
            continue
        if not (l_title or '').strip():
            errors.append(f'{_cell("Lessons", 3, row_num)}: title is required.')
            continue
        if l_type not in lesson_types:
            errors.append(f'{_cell("Lessons", 5, row_num)}: unknown lesson_type {l_type!r}.')
            continue
        minutes, ok = _as_int(minutes)
        if not ok:
            errors.append(f'{_cell("Lessons", 7, row_num)}: duration_minutes must be a whole number between 0 and 32767.')
        req_key = str(required).strip().lower() if required not in (None, '') else ''
        if req_key not in _YES_NO:
            errors.append(f'{_cell("Lessons", 8, row_num)}: required must be yes or no.')
            req_key = ''
        modules[m_order]['lessons'][order] = {
            'order': order, 'title': l_title.strip(), 'description': l_desc or '',
            'lesson_type': l_type, 'content': content or '',
            'duration_minutes': minutes, 'is_required': _YES_NO[req_key],
            'quiz_data': [],
        }

    # ── Quiz sheet (optional) ───────────────────────────────────────────
    if 'Quiz' in wb.sheetnames:
        for row_num, values in _rows(wb['Quiz']):
            values = list(values) + [None] * (len(QUIZ_HEADERS) - len(values))
            m_order, l_order, q_order, question = values[0], values[1], values[2], values[3]
            option_texts, correct = values[4:10], values[10]
            m_order, ok_m = _as_int(m_order, default=-1)
            l_order, ok_l = _as_int(l_order, default=-1)
            lesson = modules.get(m_order, {}).get('lessons', {}).get(l_order) if ok_m and ok_l else None
            if lesson is None:
                errors.append(f'{_cell("Quiz", 1, row_num)}: module_order/lesson_order do not match any lesson.')
                continue
            if lesson['lesson_type'] != Lesson.LessonType.QUIZ:
                errors.append(f'{_cell("Quiz", 2, row_num)}: lesson {m_order}/{l_order} is not a quiz.')
                continue
            if not (question or '').strip():
                errors.append(f'{_cell("Quiz", 4, row_num)}: question is required.')
                continue
            present = {
                OPTION_LETTERS[i]: str(text).strip()
                for i, text in enumerate(option_texts)
                if text not in (None, '') and str(text).strip()
            }
            if len(present) < 2:
                errors.append(f'{_cell("Quiz", 5, row_num)}: at least two options are required.')
                continue
            correct_letters = [
                c.strip().upper() for c in str(correct or '').split(',') if c.strip()
            ]
            invalid = [c for c in correct_letters if c not in present]
            if not correct_letters or invalid:
                errors.append(f'{_cell("Quiz", 11, row_num)}: correct must list letters of filled options, e.g. A,C.')
                continue
            q_order, ok = _as_int(q_order, default=len(lesson['quiz_data']) + 1)
            lesson['quiz_data'].append((q_order, {
                'question': str(question).strip(),
                'options': [
                    {'text': text, 'is_correct': letter in correct_letters}
                    for letter, text in present.items()
                ],
            }))

    if errors:
        return None, errors

    # Sort quiz questions by question_order and strip the sort key
    for module in modules.values():
        for lesson in module['lessons'].values():
            lesson['quiz_data'] = [
                q for _, q in sorted(lesson['quiz_data'], key=lambda pair: pair[0])
            ]

    course_payload['modules'] = [modules[k] for k in sorted(modules)]
    return course_payload, []
