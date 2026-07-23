from io import BytesIO

from django.contrib.auth.models import User
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from hub.models import Course, LearningPillar, Lesson, Module, UserProfile


def make_course(creator, title='Exportable Course'):
    pillar = LearningPillar.objects.get_or_create(
        slug='teach-with-ai', defaults={'name': 'Teach with AI', 'description': 'd', 'order': 1},
    )[0]
    course = Course.objects.create(
        title=title, description='Desc', pillar=pillar, level='intermediate',
        duration_hours=4, content_format='mixed',
        learning_outcomes=['Outcome one', 'Outcome two'],
        is_published=False, created_by=creator,
    )
    m1 = Module.objects.create(course=course, title='M1', description='first', order=1, duration_minutes=30)
    m2 = Module.objects.create(course=course, title='M2', order=2)
    Lesson.objects.create(module=m1, title='Intro text', lesson_type='text',
                          content='Hello', order=1, is_required=True, duration_minutes=10)
    Lesson.objects.create(module=m1, title='Watch this', lesson_type='video',
                          content='https://youtu.be/dQw4w9WgXcQ', order=2, is_required=False)
    Lesson.objects.create(
        module=m2, title='Check', lesson_type='quiz', order=1, is_required=True,
        quiz_data=[{
            'question': 'Pick two',
            'options': [
                {'text': 'A1', 'is_correct': True},
                {'text': 'B1', 'is_correct': False},
                {'text': 'C1', 'is_correct': True},
            ],
        }],
    )
    return course


class ExportXlsxTests(APITestCase):
    def setUp(self):
        self.creator = User.objects.create_user(username='xlsx_cc', password='pass12345')
        UserProfile.objects.create(user=self.creator, user_type=UserProfile.UserType.CONTENT_CREATOR)
        self.teacher = User.objects.create_user(username='xlsx_t', password='pass12345')
        UserProfile.objects.create(user=self.teacher, user_type=UserProfile.UserType.TEACHER)
        self.course = make_course(self.creator)
        self.url = reverse('authoring-course-export', kwargs={'pk': self.course.pk})

    def test_teacher_forbidden(self):
        self.client.force_authenticate(self.teacher)
        self.assertEqual(self.client.get(self.url).status_code, status.HTTP_403_FORBIDDEN)

    def test_export_workbook_structure(self):
        self.client.force_authenticate(self.creator)
        res = self.client.get(self.url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', res['Content-Type'])
        self.assertIn('exportable-course.xlsx', res['Content-Disposition'])

        wb = load_workbook(BytesIO(res.getvalue()))
        self.assertEqual(
            set(wb.sheetnames),
            {'README', 'Course', 'Modules', 'Lessons', 'Quiz', 'Choices'},
        )
        self.assertEqual(wb['Choices'].sheet_state, 'hidden')

        course_ws = wb['Course']
        self.assertEqual(course_ws['A2'].value, 'Exportable Course')
        self.assertEqual(course_ws['C2'].value, 'teach-with-ai')
        self.assertEqual(course_ws['D2'].value, 'intermediate')
        self.assertEqual(course_ws['G2'].value, 'Outcome one\nOutcome two')

        lessons_ws = wb['Lessons']
        rows = list(lessons_ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[1][4], 'video')   # lesson_type column
        self.assertEqual(rows[1][7], 'no')      # required column

        quiz_ws = wb['Quiz']
        qrow = list(quiz_ws.iter_rows(min_row=2, values_only=True))[0]
        self.assertEqual(qrow[0], 2)            # module_order (separate column)
        self.assertEqual(qrow[1], 1)            # lesson_order (separate column)
        self.assertEqual(qrow[3], 'Pick two')
        self.assertEqual(qrow[10], 'A,C')       # correct letters

    def test_export_has_dropdown_validations(self):
        self.client.force_authenticate(self.creator)
        res = self.client.get(self.url)
        wb = load_workbook(BytesIO(res.getvalue()))
        self.assertGreaterEqual(len(wb['Lessons'].data_validations.dataValidation), 2)
        self.assertGreaterEqual(len(wb['Course'].data_validations.dataValidation), 3)


class TemplateXlsxTests(APITestCase):
    def setUp(self):
        self.creator = User.objects.create_user(username='xlsx_tpl', password='pass12345')
        UserProfile.objects.create(user=self.creator, user_type=UserProfile.UserType.CONTENT_CREATOR)
        self.teacher = User.objects.create_user(username='xlsx_tpl_t', password='pass12345')
        UserProfile.objects.create(user=self.teacher, user_type=UserProfile.UserType.TEACHER)
        # A pillar must exist so the template's Choices/dropdowns have content.
        LearningPillar.objects.get_or_create(
            slug='teach-with-ai', defaults={'name': 'Teach with AI', 'description': 'd', 'order': 1},
        )
        self.url = reverse('authoring-course-template')

    def test_teacher_forbidden(self):
        self.client.force_authenticate(self.teacher)
        self.assertEqual(self.client.get(self.url).status_code, status.HTTP_403_FORBIDDEN)

    def test_template_is_blank_but_structured(self):
        self.client.force_authenticate(self.creator)
        res = self.client.get(self.url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', res['Content-Type'])
        self.assertIn('aidea-course-template.xlsx', res['Content-Disposition'])

        wb = load_workbook(BytesIO(res.getvalue()))
        self.assertEqual(
            set(wb.sheetnames),
            {'README', 'Course', 'Modules', 'Lessons', 'Quiz', 'Choices'},
        )
        # Headers present, but no data rows.
        self.assertEqual(wb['Course']['A1'].value, 'title')
        self.assertIsNone(wb['Course']['A2'].value)
        self.assertEqual(list(wb['Modules'].iter_rows(min_row=2, values_only=True)), [])
        # Dropdowns still wired up.
        self.assertGreaterEqual(len(wb['Course'].data_validations.dataValidation), 3)


class ImportXlsxTests(APITestCase):
    def setUp(self):
        self.creator = User.objects.create_user(username='xlsx_imp', password='pass12345')
        UserProfile.objects.create(user=self.creator, user_type=UserProfile.UserType.CONTENT_CREATOR)
        self.teacher = User.objects.create_user(username='xlsx_imp_t', password='pass12345')
        UserProfile.objects.create(user=self.teacher, user_type=UserProfile.UserType.TEACHER)
        self.course = make_course(self.creator, title='Round Trip')
        self.url = reverse('authoring-course-import')

    def _export_bytes(self, course):
        from io import BytesIO as B

        from hub.xlsx_transfer import build_course_workbook
        buf = B()
        build_course_workbook(course).save(buf)
        buf.seek(0)
        return buf

    def _post(self, buf, name='course.xlsx'):
        from django.core.files.uploadedfile import SimpleUploadedFile
        upload = SimpleUploadedFile(
            name, buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        return self.client.post(self.url, {'file': upload}, format='multipart')

    def test_teacher_forbidden(self):
        self.client.force_authenticate(self.teacher)
        self.assertEqual(self._post(self._export_bytes(self.course)).status_code, 403)

    def test_round_trip(self):
        self.client.force_authenticate(self.creator)
        res = self._post(self._export_bytes(self.course))
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        new = Course.objects.get(pk=res.data['id'])
        self.assertEqual(new.title, 'Round Trip (imported)')  # collision with original
        self.assertFalse(new.is_published)
        self.assertEqual(new.created_by, self.creator)
        self.assertEqual(new.level, self.course.level)
        self.assertEqual(new.learning_outcomes, self.course.learning_outcomes)
        self.assertEqual(new.modules.count(), 2)
        m2 = new.modules.get(order=2)
        quiz = m2.lessons.get(order=1)
        self.assertEqual(quiz.lesson_type, 'quiz')
        self.assertEqual(quiz.quiz_data, [{
            'question': 'Pick two',
            'options': [
                {'text': 'A1', 'is_correct': True},
                {'text': 'B1', 'is_correct': False},
                {'text': 'C1', 'is_correct': True},
            ],
        }])
        video = new.modules.get(order=1).lessons.get(order=2)
        self.assertFalse(video.is_required)

    def test_round_trip_preserves_subjects(self):
        from hub.models import Subject
        physics = Subject.objects.get(slug='physics')
        astronomy = Subject.objects.get(slug='astronomy')
        self.course.subjects.set([physics, astronomy])
        self.client.force_authenticate(self.creator)
        res = self._post(self._export_bytes(self.course))
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        new = Course.objects.get(pk=res.data['id'])
        self.assertEqual(
            set(new.subjects.values_list('slug', flat=True)), {'physics', 'astronomy'},
        )

    def test_unknown_subject_slug_rejected(self):
        from openpyxl import load_workbook
        buf = self._export_bytes(self.course)
        wb = load_workbook(buf)
        wb['Course']['H2'] = 'not-a-subject'
        from io import BytesIO as B
        out = B()
        wb.save(out)
        out.seek(0)
        self.client.force_authenticate(self.creator)
        res = self._post(out)
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(any('unknown subject' in e.lower() for e in res.data['errors']))

    def test_invalid_lesson_type_rejected_with_cell_ref(self):
        from openpyxl import load_workbook
        buf = self._export_bytes(self.course)
        wb = load_workbook(buf)
        wb['Lessons']['E2'] = 'vido'
        from io import BytesIO as B
        bad = B()
        wb.save(bad)
        bad.seek(0)
        self.client.force_authenticate(self.creator)
        res = self._post(bad)
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(any('Lessons!E2' in e for e in res.data['errors']))
        self.assertEqual(Course.objects.filter(title__startswith='Round Trip (imported').count(), 0)

    def test_wrong_extension_rejected(self):
        from io import BytesIO as B
        self.client.force_authenticate(self.creator)
        res = self._post(B(b'not a workbook'), name='course.csv')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_out_of_range_integers_rejected(self):
        # PositiveSmallIntegerField caps at 32767 on Postgres; must 400, not 500
        from io import BytesIO as B

        from openpyxl import load_workbook
        buf = self._export_bytes(self.course)
        wb = load_workbook(buf)
        wb['Course']['E2'] = 99999      # duration_hours over smallint max
        wb['Lessons']['G2'] = -5        # negative duration_minutes
        bad = B()
        wb.save(bad)
        bad.seek(0)
        self.client.force_authenticate(self.creator)
        res = self._post(bad)
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(any('Course!E2' in e for e in res.data['errors']))
        self.assertTrue(any('Lessons!G2' in e for e in res.data['errors']))
        self.assertEqual(Course.objects.filter(title__startswith='Round Trip (imported').count(), 0)

    def test_corrupt_xlsx_bytes_rejected_cleanly(self):
        # Right extension, garbage bytes — must be a clean 400, not a 500
        from io import BytesIO as B
        self.client.force_authenticate(self.creator)
        res = self._post(B(b'\x00\x01garbage not a zip'), name='course.xlsx')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('not a valid xlsx workbook', res.data['errors'][0])
