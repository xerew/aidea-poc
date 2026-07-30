from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from hub.models import (
    OnboardingDimension,
    OnboardingQuestion,
    SelfEfficacyAttempt,
    SelfEfficacyConfig,
    UserProfile,
)
from hub.models.pathway import LearningPath, UserLearningPath


def make_teacher(username='teacher1'):
    user = User.objects.create_user(username=username, password='pass')
    UserProfile.objects.create(user=user, user_type=UserProfile.UserType.TEACHER)
    return user


def make_paths():
    LearningPath.objects.create(name='Beginner Foundations', slug='beginner-foundations', competency_min=0, competency_max=2)
    LearningPath.objects.create(name='Intermediate Growth', slug='intermediate-growth', competency_min=3, competency_max=4)
    LearningPath.objects.create(name='Advanced Integration', slug='advanced-integration', competency_min=5, competency_max=6)


def all_answers(value):
    """{question_id: value} for every active question."""
    return {q.id: value for q in OnboardingQuestion.objects.filter(is_active=True)}


class SelfEfficacyGetTests(APITestCase):
    def setUp(self):
        self.user = make_teacher()
        self.client.force_authenticate(self.user)

    def test_returns_six_dimensions_with_four_questions_each(self):
        res = self.client.get(reverse('self-efficacy'))
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(len(res.data['dimensions']), 6)
        self.assertEqual(res.data['total'], 24)
        self.assertEqual(res.data['answered'], 0)
        self.assertFalse(res.data['completed'])
        for dim in res.data['dimensions']:
            self.assertEqual(len(dim['questions']), 4)
            self.assertEqual(set(dim['questions'][0].keys()), {'id', 'text'})

    def test_content_creator_can_access(self):
        creator = User.objects.create_user(username='cc', password='pass')
        UserProfile.objects.create(user=creator, user_type=UserProfile.UserType.CONTENT_CREATOR)
        self.client.force_authenticate(creator)
        self.assertEqual(self.client.get(reverse('self-efficacy')).status_code, 200)

    def test_localized_to_user_language(self):
        dim = OnboardingDimension.objects.filter(is_active=True).first()
        dim.translations = {'el': 'Γνώση ΤΝ'}
        dim.save()
        q = dim.questions.first()
        q.translations = {'el': 'Ερώτηση'}
        q.save()
        self.user.profile.language = 'el'
        self.user.profile.save()
        res = self.client.get(reverse('self-efficacy'))
        found = next(d for d in res.data['dimensions'] if d['slug'] == dim.slug)
        self.assertEqual(found['name'], 'Γνώση ΤΝ')
        self.assertEqual(found['questions'][0]['text'], 'Ερώτηση')


class SelfEfficacyPostTests(APITestCase):
    def setUp(self):
        self.user = make_teacher()
        make_paths()
        self.client.force_authenticate(self.user)

    def test_partial_save_persists_and_does_not_complete(self):
        first = OnboardingQuestion.objects.filter(is_active=True).first()
        res = self.client.post(reverse('self-efficacy'), {'answers': {str(first.id): 4}}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertFalse(res.data['completed'])
        self.assertEqual(res.data['answered'], 1)
        self.user.profile.refresh_from_db()
        self.assertEqual(self.user.profile.self_efficacy_answers[str(first.id)], 4)
        self.assertIsNone(self.user.profile.self_efficacy_completed_at)

    def test_resume_returns_saved_answers(self):
        first = OnboardingQuestion.objects.filter(is_active=True).first()
        self.client.post(reverse('self-efficacy'), {'answers': {str(first.id): 3}}, format='json')
        res = self.client.get(reverse('self-efficacy'))
        self.assertEqual(res.data['answers'][str(first.id)], 3)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_full_high_completes_and_places_advanced(self, mock_task):
        res = self.client.post(reverse('self-efficacy'), {'answers': all_answers(5)}, format='json')
        self.assertTrue(res.data['completed'])
        self.assertEqual(res.data['overall_band'], 'high')
        for dim in res.data['dimensions']:
            self.assertEqual(dim['average'], 5.0)
            self.assertEqual(dim['band'], 'high')
        self.user.profile.refresh_from_db()
        self.assertEqual(self.user.profile.competency_score, 5)  # high → 5
        self.assertIsNotNone(self.user.profile.self_efficacy_completed_at)
        path = UserLearningPath.objects.get(user=self.user)
        self.assertEqual(path.path.slug, 'advanced-integration')
        mock_task.assert_called_once_with(self.user.id)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_full_low_places_beginner(self, mock_task):
        res = self.client.post(reverse('self-efficacy'), {'answers': all_answers(1)}, format='json')
        self.assertEqual(res.data['overall_band'], 'low')
        self.user.profile.refresh_from_db()
        self.assertEqual(self.user.profile.competency_score, 1)  # low → 1

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_completing_in_two_batches(self, mock_task):
        answers = all_answers(4)
        ids = list(answers)
        half = {str(i): answers[i] for i in ids[:12]}
        rest = {str(i): answers[i] for i in ids[12:]}
        r1 = self.client.post(reverse('self-efficacy'), {'answers': half}, format='json')
        self.assertFalse(r1.data['completed'])
        r2 = self.client.post(reverse('self-efficacy'), {'answers': rest}, format='json')
        self.assertTrue(r2.data['completed'])
        self.assertEqual(r2.data['overall_band'], 'high')  # 4.0 → high

    def test_rating_out_of_range_rejected(self):
        first = OnboardingQuestion.objects.filter(is_active=True).first()
        res = self.client.post(reverse('self-efficacy'), {'answers': {str(first.id): 6}}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_unknown_question_rejected(self):
        res = self.client.post(reverse('self-efficacy'), {'answers': {'999999': 3}}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_content_creator_completes_without_learning_path(self):
        creator = User.objects.create_user(username='cc2', password='pass')
        UserProfile.objects.create(user=creator, user_type=UserProfile.UserType.CONTENT_CREATOR)
        self.client.force_authenticate(creator)
        res = self.client.post(reverse('self-efficacy'), {'answers': all_answers(5)}, format='json')
        self.assertTrue(res.data['completed'])
        creator.profile.refresh_from_db()
        self.assertEqual(creator.profile.competency_score, 5)   # recorded
        self.assertEqual(SelfEfficacyAttempt.objects.filter(user=creator).count(), 1)
        self.assertFalse(UserLearningPath.objects.filter(user=creator).exists())  # no placement


class SelfEfficacyAttemptAndRetakeTests(APITestCase):
    def setUp(self):
        self.user = make_teacher()
        make_paths()
        self.admin = User.objects.create_user(username='se_admin', password='pass')
        UserProfile.objects.create(user=self.admin, user_type=UserProfile.UserType.ADMIN)
        self.client.force_authenticate(self.user)

    def _complete(self, value=4):
        return self.client.post(reverse('self-efficacy'), {'answers': all_answers(value)}, format='json')

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_completion_snapshots_an_attempt(self, _m):
        self._complete(4)
        attempts = SelfEfficacyAttempt.objects.filter(user=self.user)
        self.assertEqual(attempts.count(), 1)
        a = attempts.first()
        self.assertEqual(a.overall_band, 'high')       # 4.0
        self.assertEqual(a.competency_score, 5)
        self.assertEqual(len(a.answers), 24)
        self.assertIn('ai-knowledge', a.dimension_scores)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_completed_assessment_is_read_only(self, _m):
        self._complete(4)
        # Any further submit is rejected until an admin opens a retake.
        first = OnboardingQuestion.objects.filter(is_active=True).first()
        res = self.client.post(reverse('self-efficacy'), {'answers': {str(first.id): 1}}, format='json')
        self.assertEqual(res.status_code, 400)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_retake_blocked_when_closed(self, _m):
        self._complete(4)
        res = self.client.post(reverse('self-efficacy-retake'))
        self.assertEqual(res.status_code, 403)

    def _open_window(self):
        cfg = SelfEfficacyConfig.get()
        cfg.retake_open = True
        cfg.retake_opened_at = timezone.now()
        cfg.save()

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_retake_when_open_keeps_history(self, _m):
        self._complete(4)  # attempt 1: high
        self._open_window()

        res = self.client.get(reverse('self-efficacy'))
        self.assertTrue(res.data['can_retake'])

        retake = self.client.post(reverse('self-efficacy-retake'))
        self.assertEqual(retake.status_code, 200)
        self.assertTrue(retake.data['completed'])    # old review kept
        self.assertTrue(retake.data['retaking'])
        self.assertEqual(retake.data['answers'], {})  # fresh draft

        self._complete(1)  # attempt 2: low
        self.assertEqual(SelfEfficacyAttempt.objects.filter(user=self.user).count(), 2)
        self.user.profile.refresh_from_db()
        self.assertEqual(self.user.profile.competency_score, 1)  # re-placed from new attempt
        self.assertFalse(self.user.profile.self_efficacy_retaking)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_paused_retake_keeps_old_review(self, _m):
        self._complete(4)  # attempt 1: high, competency 5
        self._open_window()
        self.client.post(reverse('self-efficacy-retake'))  # start

        # Answer a few, then walk away (no finish).
        ids = list(all_answers(1))
        partial = {str(ids[0]): 1, str(ids[1]): 1, str(ids[2]): 1}
        self.client.post(reverse('self-efficacy'), {'answers': partial}, format='json')

        res = self.client.get(reverse('self-efficacy'))
        self.assertTrue(res.data['completed'])         # still completed
        self.assertTrue(res.data['retaking'])
        self.assertEqual(res.data['overall_band'], 'high')  # OLD review, not the draft
        self.assertEqual(len(res.data['answers']), 3)  # draft resumes where left off
        self.user.profile.refresh_from_db()
        self.assertEqual(self.user.profile.competency_score, 5)  # unchanged until finish
        self.assertEqual(SelfEfficacyAttempt.objects.filter(user=self.user).count(), 1)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_only_one_retake_per_window(self, _m):
        self._complete(4)
        self._open_window()
        self.client.post(reverse('self-efficacy-retake'))
        self._complete(3)  # second attempt done within the window
        # Window is still open, but this teacher has used their one retake.
        res = self.client.get(reverse('self-efficacy'))
        self.assertFalse(res.data['can_retake'])
        again = self.client.post(reverse('self-efficacy-retake'))
        self.assertEqual(again.status_code, 403)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_new_window_allows_another_retake(self, _m):
        self._complete(4)
        self._open_window()
        self.client.post(reverse('self-efficacy-retake'))
        self._complete(3)
        # Admin opens a fresh window → the teacher may retake once more.
        self._open_window()
        res = self.client.get(reverse('self-efficacy'))
        self.assertTrue(res.data['can_retake'])

    def test_retake_without_completion_rejected(self):
        self._open_window()
        res = self.client.post(reverse('self-efficacy-retake'))
        self.assertEqual(res.status_code, 400)


class AdminSelfEfficacyTests(APITestCase):
    def setUp(self):
        self.teacher = make_teacher()
        make_paths()
        self.admin = User.objects.create_user(username='se_admin2', password='pass')
        UserProfile.objects.create(user=self.admin, user_type=UserProfile.UserType.ADMIN)

    def test_toggle_requires_admin(self):
        self.client.force_authenticate(self.teacher)
        self.assertEqual(self.client.get(reverse('admin-self-efficacy')).status_code, 403)

    def test_admin_toggles_retake_records_opened_at(self):
        self.client.force_authenticate(self.admin)
        got = self.client.get(reverse('admin-self-efficacy')).data
        self.assertFalse(got['retake_open'])
        self.assertIsNone(got['retake_opened_at'])
        res = self.client.patch(reverse('admin-self-efficacy'), {'retake_open': True}, format='json')
        self.assertTrue(res.data['retake_open'])
        self.assertIsNotNone(res.data['retake_opened_at'])  # window timestamp set
        self.assertTrue(SelfEfficacyConfig.get().retake_open)

    @patch('hub.tasks.compute_user_recommendations.delay')
    def test_export_has_versioned_sheets(self, _m):
        # Teacher completes, admin opens a window, teacher retakes → 2 attempts.
        self.client.force_authenticate(self.teacher)
        self.client.post(reverse('self-efficacy'), {'answers': all_answers(5)}, format='json')
        cfg = SelfEfficacyConfig.get()
        cfg.retake_open = True
        cfg.retake_opened_at = timezone.now()
        cfg.save()
        self.client.post(reverse('self-efficacy-retake'))
        self.client.post(reverse('self-efficacy'), {'answers': all_answers(2)}, format='json')

        self.client.force_authenticate(self.admin)
        res = self.client.get(reverse('admin-self-efficacy-export'))
        self.assertEqual(res.status_code, 200)
        self.assertIn('spreadsheetml', res['Content-Type'])

        wb = load_workbook(BytesIO(res.content))
        self.assertIn('AI Comp Version 1', wb.sheetnames)
        self.assertIn('AI Comp Version 2', wb.sheetnames)
        self.assertIn('Questions', wb.sheetnames)
        ws = wb['AI Comp Version 1']
        header = [c.value for c in ws[1]]
        self.assertEqual(header[:5], ['User ID', 'Name', 'Username', 'Email', 'Date'])
        self.assertIn('Q24', header)
        self.assertIn('Competency score (0-6)', header)
        # One data row for our single teacher on version 1.
        self.assertEqual(ws.max_row, 2)
        # The Questions reference sheet lists all 24 questions.
        self.assertEqual(wb['Questions'].max_row, 25)  # header + 24

    def test_export_requires_admin(self):
        self.client.force_authenticate(self.teacher)
        self.assertEqual(self.client.get(reverse('admin-self-efficacy-export')).status_code, 403)
